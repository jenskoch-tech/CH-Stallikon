import requests
from bs4 import BeautifulSoup
import csv
import re
import time
import concurrent.futures
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

BASE_URL = "https://explorer-magazin.com/anzeige/page/1/"
HEADERS = {"User-Agent": "Mozilla/5.0"}  # Avoid blocking

# Store discovered URLs for listings
urls_to_visit = [BASE_URL]
all_listings = []
visited_pages = set()

# Regex to match pagination links
pagination_pattern = re.compile(r"/anzeige/page/\d+/")

# Limit the number of pages to prevent infinite loops
MAX_PAGES = 40  # Set a reasonable limit
MAX_LISTINGS = 900  # Prevent excessive data collection
REQUEST_TIMEOUT = 15  # Timeout for requests

def scrape_listings():
    page_count = 0
    while urls_to_visit and page_count < MAX_PAGES and len(all_listings) < MAX_LISTINGS:
        current_url = urls_to_visit.pop(0)
        if current_url in visited_pages:
            continue
        
        try:
            response = requests.get(current_url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
        except requests.RequestException as e:
            print(f"Failed to fetch {current_url}: {e}")
            continue

        soup = BeautifulSoup(response.content, "html.parser")
        
        # Find all individual listing links
        for link in soup.find_all("a", href=True):
            url = link["href"]
            if url.startswith("https://explorer-magazin.com/anzeige/") and url not in all_listings and "/page/" not in url:
                all_listings.append(url)
                if len(all_listings) >= MAX_LISTINGS:
                    break
        
        # Find pagination links
        for page_link in soup.find_all("a", href=True):
            page_url = page_link["href"]
            if pagination_pattern.search(page_url) and page_url not in urls_to_visit:
                urls_to_visit.append(page_url)
        
        visited_pages.add(current_url)
        page_count += 1
        print(f"Scraped page {page_count}: {current_url}")
        time.sleep(1)  # Prevent excessive requests

def scrape_vehicle_details(url):
    try:
        response = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"Error scraping {url}: {e}")
        return None
    
    soup = BeautifulSoup(response.content, "html.parser")
    
    try:
        name_element = soup.select_one("span.wpa-row-name--fahrzeugbeschreibung")
        price_element = None
        mileage_element = soup.select_one("span.wpa-row-name--kilometerstand")
        year_element = soup.select_one("span.wpa-row-name--baujahr_dd")
        type_element = soup.select_one("span.wpa-row-name--art_dropdown")
        
        # Look for price span based on inline style
        for span in soup.find_all("span"):
            if span.has_attr("style") and "background-color:#ab0000" in span["style"]:
                price_element = span
                break
        
        name = name_element.text.strip() if name_element else "Unknown"
        price = price_element.text.strip() if price_element else "Not Listed"
        mileage = mileage_element.text.strip() if mileage_element else "Unknown"
        year = year_element.text.strip() if year_element else "Unknown"
        vehicle_type = type_element.text.strip() if type_element else "Unknown"
        
        # Clean price formatting (convert to numeric format for sorting)
        price = re.sub(r"[^0-9]", "", price)
        price = int(price) if price.isdigit() else None
    except AttributeError:
        return None
    
    return {"Name": name, "Price": price, "Mileage": mileage, "Year": year, "Type": vehicle_type, "URL": url}

def save_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicles"
    
    headers = ["Name", "Price", "Mileage", "Year", "Type", "URL"]
    ws.append(headers)
    
    for row in data:
        row["URL"] = f'=HYPERLINK("{row["URL"]}", "Link")'
        ws.append([row[col] for col in headers])
    
    # Apply auto filter
    ws.auto_filter.ref = ws.dimensions
    
    # Adjust column width
    for col_num, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    wb.save("vehicles.xlsx")

# Run scraping steps
scrape_listings()
vehicle_data = []

# Use threading to speed up vehicle detail scraping
with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
    results = executor.map(scrape_vehicle_details, all_listings)
    vehicle_data = [res for res in results if res]

save_to_excel(vehicle_data)
print("Scraping complete! Data saved to vehicles.xlsx")
